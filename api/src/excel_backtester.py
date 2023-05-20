from datetime import datetime
from typing import Any, Optional, Tuple, Union
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
from pycel import ExcelCompiler
from api.src.worksheet_utils import CellPostion, format_coordinate, get_positioned_cell, parse_coordinate
from nersent_pace_py import nersent_pace_py as pace


def format_config_column_id(column_id: str) -> str:
    return f"<nersent_pace::config::{column_id}>"


def format_data_column_id(column_id: str) -> str:
    return f"<nersent_pace::data::{column_id}>"


def format_input_column_id(column_id: str) -> str:
    return f"<nersent_pace::input::{column_id}>"


def format_output_column_id(column_id: str) -> str:
    return f"<nersent_pace::output::{column_id}>"


def format_stats_column_id(column_id: str) -> str:
    return f"<nersent_pace::metrics::{column_id}>"


def format_target_annotation(target: str) -> str:
    return f"<nersent_pace::target::{target}>"


def get_cell_position(value: str) -> CellPostion:
    if value is None:
        return CellPostion.Bottom

    value = value.lower()

    if format_target_annotation("right") in value:
        return CellPostion.Right

    if format_target_annotation("left") in value:
        return CellPostion.Left

    if format_target_annotation("top") in value:
        return CellPostion.Top

    if format_target_annotation("bottom") in value:
        return CellPostion.Bottom

    return CellPostion.Bottom


class ExcelBacktester():
    def __init__(self):
        self.worksheet_name: Optional[str] = None
        self.path: Optional[str] = None
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.compiled_worksheet: Optional[ExcelCompiler] = None
        self.is_loaded: bool = False

        config_columns = [
            "on_bar_close",
            "initial_capital",
            "buy_with_equity",
            "risk_free_rate"
        ]

        data_columns = [
            "time",
            "open",
            "high",
            "low",
            "close",
            "volume"
        ]

        input_columns = [
            "strategy_signal"
        ]

        output_columns = [
            "time",
            "tick",
            "equity",
            "net_equity",
            "open_profit",
            "position_size",
            "returns",
            "direction",
            "logs",
            "pinescript",
            "omega_ratio",
            "sharpe_ratio",
            "sortino_ratio",
            "profitable",
            "max_drawdown",
            "max_drawdown_pct",
            "max_run_up",
            "max_run_up_pct",
            "net_profit",
            "net_profit_pct",
            "gross_profit",
            "gross_profit_pct",
            "gross_loss",
            "gross_loss_pct",
            "closed_trades",
            "winning_trades",
            "losing_trades",
            "profit_factor",
            "equity_curve_max_drawdown_pct",
            "intra_trade_max_drawdown_pct",
            "net_profit_l_s_ratio"
        ]

        stats_columns = [*output_columns]

        self.config_columns = list(
            map(format_config_column_id, config_columns))
        self.data_columns = list(map(format_data_column_id, data_columns))
        self.input_columns = list(
            map(format_input_column_id, input_columns))
        self.output_columns = list(
            map(format_output_column_id, output_columns))
        self.stats_columns = list(
            map(format_stats_column_id, stats_columns))

        self.columns = [
            *self.config_columns,
            *self.data_columns,
            *self.input_columns,
            *self.output_columns,
            *self.stats_columns
        ]

    def load(self, path: str, worksheet_name: str):
        self.is_loaded = False

        self.worksheet_name = worksheet_name
        self.path = path
        xlxs_path = os.path.abspath(path)
        self.workbook = load_workbook(filename=path)
        self.worksheet = self.workbook[worksheet_name]
        self.compiled_worksheet = ExcelCompiler(filename=xlxs_path)

        self.is_loaded = True

    def compute(self) -> dict[str, Union[str, int, float]]:
        if not self.is_loaded:
            raise Exception("ExcelBacktester is not loaded")

        column_coordinate_map: dict[str, Tuple[str, int]] = {}

        for row in self.worksheet.iter_rows():
            for cell in row:
                for pace_column_id in self.columns:
                    if cell is None or cell.value is None:
                        continue
                    value = str(cell.value).lower()
                    if pace_column_id in value:
                        coordinate = cell.coordinate
                        parsed_coordinate = parse_coordinate(coordinate)
                        column_coordinate_map[pace_column_id] = parsed_coordinate

        assert_columns = [*self.data_columns, *self.input_columns]

        for column_id in assert_columns:
            if column_id not in column_coordinate_map:
                raise Exception(
                    f"Could not find data column {column_id} in worksheet")

        print("Column coordinates:")
        print(column_coordinate_map)

        column_to_length_map: dict[str, int] = {}

        for column_id in self.data_columns:
            (column, row) = column_coordinate_map[column_id]

            for i in range(row, self.worksheet.max_row + 1):
                cell = self.worksheet[column + str(i)]
                if cell is None or cell.value is None:
                    print(f"[{column_id}]: {i}")
                    break
                column_to_length_map[column_id] = i

        print("Column lengths:")
        print(column_to_length_map)

        data_length: Optional[int] = None

        for column_id in self.data_columns:
            column_length = column_to_length_map[column_id]
            if data_length is None:
                data_length = column_length
                continue
            if column_length < data_length:
                data_length = column_length

        data_length = data_length - 1
        data: dict[str, list[float]] = {}

        for column_id in self.data_columns:
            (column, row) = column_coordinate_map[column_id]
            data[column_id] = []

            for i in range(row + 1, data_length + 2):
                cell = self.worksheet[column + str(i)]

                if cell is None or cell.value is None:
                    break

                if isinstance(cell.value, datetime):
                    data[column_id].append(cell.value.timestamp())
                else:
                    data[column_id].append(float(cell.value))

        # print(data_length)

        for column_id in data:
            data[column_id] = data[column_id][:data_length]
            # if len(data[column_id]) != data_length:
            #     raise Exception(
            #         f"Data column {column_id} does not have the same length as other data columns: {len(data[column_id])} vs {data_length} | {data[column_id][0]} | {data[column_id][-1]}")

        (signal_column, signal_row) = column_coordinate_map[format_input_column_id(
            "strategy_signal")]
        signal_row_start = signal_row + 1
        signal_row_end = signal_row + data_length

        evaluation_target = f"{self.worksheet_name}!{signal_column}{signal_row_start}:{signal_column}{signal_row_end}"

        evaluated_signals = self.compiled_worksheet.evaluate(evaluation_target)

        if len(evaluated_signals) != data_length:
            raise Exception(
                f"Evaluated signals length does not match data length: {len(evaluated_signals)} vs {data_length}")

        signals: list[pace.StrategySignal] = []

        for i in range(len(evaluated_signals)):
            signal: str = evaluated_signals[i]
            pace_signal_id = "hold"

            if signal is not None:
                signal = signal.lower().strip()
                if signal == "long":
                    pace_signal_id = "long"
                elif signal == "long_entry" or signal == "long entry":
                    pace_signal_id = "long_entry"
                elif signal == "long_exit" or signal == "long exit":
                    pace_signal_id = "long_exit"
                elif signal == "short":
                    pace_signal_id = "short"
                elif signal == "short_entry" or signal == "short entry":
                    pace_signal_id = "short_entry"
                elif signal == "short_exit" or signal == "short exit":
                    pace_signal_id = "short_exit"

            signal = pace.StrategySignal({"id": pace_signal_id})
            signals.append(signal)

        assert len(signals) == data_length

        data_provider = pace.DataProvider({
            "time": data[format_data_column_id("time")],
            "open": data[format_data_column_id("open")],
            "close": data[format_data_column_id("close")],
            "high": data[format_data_column_id("high")],
            "low": data[format_data_column_id("low")],
            "volume": data[format_data_column_id("volume")],
        })

        on_bar_close_column = format_config_column_id("on_bar_close")
        initial_capital_column = format_config_column_id("initial_capital")
        buy_with_equity_column = format_config_column_id("buy_with_equity")
        risk_free_rate_column = format_config_column_id("risk_free_rate")

        backtest_config = {
            "on_bar_close": False,
            "initial_capital": 1000.0,
            "buy_with_equity": False,
            "risk_free_rate": 0.0,
        }

        if on_bar_close_column in column_coordinate_map:
            coordinate = column_coordinate_map[on_bar_close_column]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_cell_coordinate = format_coordinate(target_cell)
            target_cell_value = self.worksheet[target_cell_coordinate].value

            if target_cell_value is not None:
                backtest_config["on_bar_close"] = int(target_cell_value) == 1

        if initial_capital_column in column_coordinate_map:
            coordinate = column_coordinate_map[initial_capital_column]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_cell_coordinate = format_coordinate(target_cell)
            target_cell_value = self.worksheet[target_cell_coordinate].value

            if target_cell_value is not None:
                backtest_config["initial_capital"] = float(target_cell_value)

        if buy_with_equity_column in column_coordinate_map:
            coordinate = column_coordinate_map[buy_with_equity_column]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_cell_coordinate = format_coordinate(target_cell)
            target_cell_value = self.worksheet[target_cell_coordinate].value

            if target_cell_value is not None:
                backtest_config["buy_with_equity"] = int(
                    target_cell_value) == 1

        if risk_free_rate_column in column_coordinate_map:
            coordinate = column_coordinate_map[risk_free_rate_column]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_cell_coordinate = format_coordinate(target_cell)
            target_cell_value = self.worksheet[target_cell_coordinate].value

            if target_cell_value is not None:
                backtest_config["risk_free_rate"] = float(target_cell_value)

        print(backtest_config)

        backtest_res = pace.run_backtest(
            data_provider, backtest_config, signals)

        column_update_map: dict[str, Union[str, int, float]] = {}

        def unwrap_bar_value(id: str, bar) -> Optional[Any]:
            if id is None:
                raise Exception("id is None")
            if bar is None:
                return None

            target_ids = {
                "tick": bar.tick,
                "time": bar.time,
                "equity": bar.equity,
                "net_equity": bar.net_equity,
                "open_profit": bar.open_profit,
                "position_size": bar.position_size,
                "returns": bar.returns,
                "direction": bar.direction,
                "logs": bar.logs,
                "omega_ratio": bar.omega_ratio,
                "sharpe_ratio": bar.sharpe_ratio,
                "sortino_ratio": bar.sortino_ratio,
                "profitable": bar.profitable,
                "max_drawdown": bar.max_drawdown,
                "max_drawdown_pct": bar.max_drawdown_percent,
                "max_run_up": bar.max_run_up,
                "max_run_up_pct": bar.max_run_up_percent,
                "net_profit": bar.net_profit,
                "net_profit_pct": bar.net_profit_percent,
                "gross_profit": bar.gross_profit,
                "gross_profit_pct": bar.gross_profit_percent,
                "gross_loss": bar.gross_loss,
                "gross_loss_pct": bar.gross_loss_percent,
                "closed_trades": bar.closed_trades,
                "winning_trades": bar.winning_trades,
                "losing_trades": bar.losing_trades,
                "profit_factor": bar.profit_factor,
                "equity_curve_max_drawdown_pct": bar.equity_curve_max_drawdown,
                "intra_trade_max_drawdown_pct": bar.intra_trade_max_drawdown,
                "net_profit_l_s_ratio": bar.net_profit_l_s_ratio
            }

            for (target_id, value) in target_ids.items():
                if format_output_column_id(target_id) == id or format_stats_column_id(target_id) == id:
                    return value

        for (i, bar) in enumerate(backtest_res.bars):
            for id in self.output_columns:
                if id not in column_coordinate_map:
                    continue

                (column, start_row) = column_coordinate_map[id]
                target_coordinate = f"{column}{start_row + i + 1}"

                value = unwrap_bar_value(id, bar)
                column_update_map[target_coordinate] = value

        print("Computed data columns")

        # iterate over stats_columns
        for stat_column_id in self.stats_columns:
            if stat_column_id not in column_coordinate_map:
                continue

            coordinate = column_coordinate_map[stat_column_id]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_coordinate = format_coordinate(target_cell)

            last_bar = backtest_res.bars[-1]

            value = unwrap_bar_value(stat_column_id, last_bar)
            if value is None:
                value = ""

            column_update_map[target_coordinate] = value

        print("Computed stast columns")

        pinescript_column = format_stats_column_id("pinescript")

        if pinescript_column in column_coordinate_map:
            coordinate = column_coordinate_map[pinescript_column]
            value = self.worksheet[format_coordinate(coordinate)]
            target_position = get_cell_position(value.value)
            target_cell = get_positioned_cell(coordinate, target_position)
            target_coordinate = format_coordinate(target_cell)

            column_update_map[target_coordinate] = backtest_res.pinescript

        print("Computed PineScript column")

        return column_update_map
