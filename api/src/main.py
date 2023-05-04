import json
import os
from time import time
from typing import Optional, Tuple, Union
from nersent_pace_py import nersent_pace_py as pace
import pandas as pd
from pycel import ExcelCompiler
from openpyxl import load_workbook
import gspread
from gspread.utils import ExportFormat
from gspread import Cell
from api.src.excel_backtester import ExcelBacktester
from api.src.worksheet_utils import format_coordinate, get_cell_coordinate_below, parse_coordinate

if __name__ == "__main__":
    gc = gspread.service_account(
        filename=os.path.abspath("service_account.json")
    )
    google_sheet = gc.open_by_url(
        "https://docs.google.com/spreadsheets/d/1FPMGmFUMX_yeMHHalOywVxPk3WrmWgHZ6H4lC6thQRg/edit?usp=sharing")

    # print(sheet.update_title("aha"))
    raw_exported_google_sheet = google_sheet.export(ExportFormat.EXCEL)
    timestamp = str(time())
    xlxs_path = os.path.abspath(f"local/fixtures/exported_{timestamp}.xlsx")

    with open(xlxs_path, "wb") as f:
        f.write(raw_exported_google_sheet)

    worksheet_name = "btc_1d"
    google_worksheet = google_sheet.worksheet(worksheet_name)

    backtester = ExcelBacktester()
    backtester.load(xlxs_path, worksheet_name)
    update_map = backtester.compute()

    google_worksheet_cells_to_update: list[Cell] = []

    # save update_map to json at path os.path.abspath("local/fixtures/update_map.json")
    with open(os.path.abspath("local/fixtures/update_map.json"), "w") as f:
        f.write(json.dumps(update_map))

    google_worksheet_batch: list[dict] = []

    for coordinate, value in update_map.items():
        google_worksheet_batch.append({
            "range": coordinate,
            "values": [[value]]
        })

    google_worksheet.batch_update(google_worksheet_batch)


if False:
    xlxs_path = os.path.abspath("local/fixtures/btc_1d.xlsx")
    workbook = load_workbook(filename=xlxs_path)
    worksheet = workbook.active
    worksheet_name = worksheet.title
    compiled_excel = ExcelCompiler(filename=xlxs_path)

    def format_config_column_id(column_id: str) -> str:
        return f"<nersent_pace::config::{column_id}>"

    def format_data_column_id(column_id: str) -> str:
        return f"<nersent_pace::data::{column_id}>"

    def format_input_column_id(column_id: str) -> str:
        return f"<nersent_pace::input::{column_id}>"

    def format_output_column_id(column_id: str) -> str:
        return f"<nersent_pace::output::{column_id}>"

    config_columns = [
        "on_bar_close",
        "initial_capital",
        "buy_with_equity",
        "risk_free_rate"
    ]

    data_columns = [
        "time",
        "open",
        "high",
        "low",
        "close",
        "volume"
    ]
    input_columns = [
        "strategy_signal"
    ]
    output_columns = [
        "time",
        "tick",
        "equity",
        "net_equity",
        "open_profit",
        "position_size",
        "returns",
        "direction",
        "logs",
        "pinescript"
    ]

    config_columns = list(map(format_config_column_id, config_columns))
    data_columns = list(map(format_data_column_id, data_columns))
    input_columns = list(
        map(format_input_column_id, input_columns))
    output_columns = list(map(format_output_column_id, output_columns))

    columns = [
        *config_columns,
        *data_columns,
        *input_columns,
        *output_columns
    ]

    # given pace_column_ids, find all the column indexes in worksheet by checking if cell text contains the column id
    column_coordinate_map: dict[str, Tuple[str, int]] = {}

    for row in worksheet.iter_rows():
        for cell in row:
            for pace_column_id in columns:
                if cell is None or cell.value is None:
                    continue
                value = str(cell.value).lower()
                if pace_column_id in value:
                    coordinate = cell.coordinate
                    parsed_coordinate = parse_coordinate(coordinate)
                    column_coordinate_map[pace_column_id] = parsed_coordinate

    assert_columns = [*data_columns, *input_columns]

    for column_id in assert_columns:
        if column_id not in column_coordinate_map:
            raise Exception(
                f"Could not find data column {column_id} in worksheet")

    column_to_last_row_index_map: dict[str, int] = {}
    column_to_length_map: dict[str, int] = {}

    for column_id in data_columns:
        (column, row) = column_coordinate_map[column_id]

        column_to_last_row_index_map[column_id] = -1

        for i in range(row, worksheet.max_row + 1):
            cell = worksheet[column + str(i)]
            if cell is None or cell.value is None:
                break
            column_to_last_row_index_map[column_id] = i
            column_to_length_map[column_id] = i - row + 1

    data_length: Optional[int] = None

    for column_id in data_columns:
        column_length = column_to_length_map[column_id]
        if data_length is None:
            data_length = column_length
            continue
        if column_length < data_length:
            data_length = column_length

    data_length = data_length - 1
    data: dict[str, list[float]] = {}

    for column_id in data_columns:
        (column, row) = column_coordinate_map[column_id]
        data[column_id] = []

        for i in range(row + 1, data_length + 2):
            cell = worksheet[column + str(i)]
            if cell is None or cell.value is None:
                break
            data[column_id].append(int(cell.value))

    for column_id in data:
        if len(data[column_id]) != data_length:
            raise Exception(
                f"Data column {column_id} does not have the same length as other data columns: {len(data[column_id])} vs {data_length} | {data[column_id][0]} | {data[column_id][-1]}")

    (signal_column, signal_row) = column_coordinate_map[format_input_column_id(
        "strategy_signal")]
    signal_row_start = signal_row + 1
    signal_row_end = signal_row + data_length

    evaluation_target = f"{worksheet_name}!{signal_column}{signal_row_start}:{signal_column}{signal_row_end}"

    evaluated_signals = compiled_excel.evaluate(evaluation_target)

    if len(evaluated_signals) != data_length:
        raise Exception(
            f"Evaluated signals length does not match data length: {len(evaluated_signals)} vs {data_length}")

    signals: list[pace.StrategySignal] = []

    for i in range(len(evaluated_signals)):
        signal: str = evaluated_signals[i]
        pace_signal_id = "hold"

        if signal is not None:
            signal = signal.lower().strip()
            if signal == "long":
                pace_signal_id = "long"
            elif signal == "long_entry" or signal == "long entry":
                pace_signal_id = "long_entry"
            elif signal == "long_exit" or signal == "long exit":
                pace_signal_id = "long_exit"
            elif signal == "short":
                pace_signal_id = "short"
            elif signal == "short_entry" or signal == "short entry":
                pace_signal_id = "short_entry"
            elif signal == "short_exit" or signal == "short exit":
                pace_signal_id = "short_exit"

        signal = pace.StrategySignal({"id": pace_signal_id})
        signals.append(signal)

    assert len(signals) == data_length

    data_provider = pace.DataProvider({
        "time": data[format_data_column_id("time")],
        "open": data[format_data_column_id("open")],
        "close": data[format_data_column_id("close")],
        "high": data[format_data_column_id("high")],
        "low": data[format_data_column_id("low")],
        "volume": data[format_data_column_id("volume")],
    })

    on_bar_close_column = format_config_column_id("on_bar_close")
    initial_capital_column = format_config_column_id("initial_capital")
    buy_with_equity_column = format_config_column_id("buy_with_equity")
    risk_free_rate_column = format_config_column_id("risk_free_rate")

    backtest_config = {
        "on_bar_close": False,
        "initial_capital": 1000.0,
        "buy_with_equity": True,
        "risk_free_rate": 0.0,
    }

    if on_bar_close_column in column_coordinate_map:
        coordinate = format_coordinate(get_cell_coordinate_below(
            column_coordinate_map[on_bar_close_column]))
        backtest_config["on_bar_close"] = int(worksheet[coordinate].value) == 1

    if initial_capital_column in column_coordinate_map:
        coordinate = format_coordinate(get_cell_coordinate_below(
            column_coordinate_map[initial_capital_column]))
        backtest_config["initial_capital"] = float(worksheet[coordinate].value)

    if buy_with_equity_column in column_coordinate_map:
        coordinate = format_coordinate(get_cell_coordinate_below(
            column_coordinate_map[buy_with_equity_column]))
        backtest_config["buy_with_equity"] = int(
            worksheet[coordinate].value) == 1

    if risk_free_rate_column in column_coordinate_map:
        coordinate = format_coordinate(get_cell_coordinate_below(
            column_coordinate_map[risk_free_rate_column]))
        backtest_config["risk_free_rate"] = float(worksheet[coordinate].value)

    backtest_res = pace.run_backtest(data_provider, backtest_config, signals)

    column_update_map: dict[str, Union[float, str]] = {}

    for (i, bar) in enumerate(backtest_res.bars):
        for id in output_columns:
            if id not in column_coordinate_map:
                continue

            (column, start_row) = column_coordinate_map[id]
            target_coordinate = f"{column}{start_row + i + 1}"

            if id == format_output_column_id("tick"):
                column_update_map[target_coordinate] = bar.tick
            elif id == format_output_column_id("time"):
                column_update_map[target_coordinate] = bar.time
            elif id == format_output_column_id("equity"):
                column_update_map[target_coordinate] = bar.equity
            elif id == format_output_column_id("net_equity"):
                column_update_map[target_coordinate] = bar.net_equity
            elif id == format_output_column_id("open_profit"):
                column_update_map[target_coordinate] = bar.open_profit
            elif id == format_output_column_id("position_size"):
                column_update_map[target_coordinate] = bar.position_size
            elif id == format_output_column_id("returns"):
                column_update_map[target_coordinate] = bar.returns
            elif id == format_output_column_id("direction"):
                column_update_map[target_coordinate] = bar.direction
            elif id == format_output_column_id("logs"):
                column_update_map[target_coordinate] = bar.logs

    pinescript_column = format_output_column_id("pinescript")

    if pinescript_column in column_coordinate_map:
        target_coordinate = get_cell_coordinate_below(
            column_coordinate_map[pinescript_column])
        target_coordinate = format_coordinate(target_coordinate)
        column_update_map[target_coordinate] = backtest_res.pinescript

    update_original_file = True

    if update_original_file:
        for coordinate in column_update_map:
            worksheet[coordinate] = column_update_map[coordinate]

        workbook.save(filename=xlxs_path)
